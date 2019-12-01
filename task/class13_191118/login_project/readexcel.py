"""
============================
Author  : XiaoLei.Du
Time    : 2019/11/18 21:04
E-mail  : 506615839@qq.com
Software: PyCharm
============================
"""
'''
为什么要封装？
为了使用更加方便，提高代码重用率
封装的需求是什么？
读取数据的方法
'''
import openpyxl

class ReadExcel(object):

    def __init__(self,filename,sheetname):
        '''
        初始化方法
        :param filename: 文件名
        :param sheetname: 表单名
        '''
        self.filename=filename
        self.sheetname=sheetname

    def open(self):
        '''打开工作簿，选中表单'''
        self.wb=openpyxl.load_workbook(self.filename)
        self.sh=self.wb[self.sheetname]

    def save(self):
        '''保存工作簿对象方法'''
        self.wb.save(self.filename)
        self.wb.close() # 这一行加不加关系不大，加了可以释放内存

    def read_data(self):
        '''读取数据'''
        # 1、打开工作簿，选中表单
        self.open()
        # 获取最大行
        max_row=self.sh.max_row

        # 读取所有的数据，放到一个列表中
        list_data=[]
        for i in range(1,max_row+1):
            data1 = self.sh.cell(row=i, column=1).value
            data2 = self.sh.cell(row=i, column=2).value
            data3 = self.sh.cell(row=i, column=3).value
            data4 = self.sh.cell(row=i, column=4).value
            list_data.append([data1,data2,data3,data4])
        cases=[]
        # 获取第一行数据表头
        title=list_data[0]
        for item in list_data[1:]:
            # 遍历第一行之外的其他数据，聚合打包成字典
            case=dict(zip(title,item))
            cases.append(case)
        return cases



if __name__ =="__main__":
    excel = ReadExcel('case.xlsx', 'login')
    print(excel.read_data())

