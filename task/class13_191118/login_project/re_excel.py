"""
============================
Author  : XiaoLei.Du
Time    : 2019/11/19 7:02
E-mail  : 506615839@qq.com
Software: PyCharm
============================
"""
import openpyxl
# 第一步：打开工作簿对象
workbook=openpyxl.load_workbook('case.xlsx')
# 第二步：选中表单
sheet=workbook['login']
# print(list(sheet))
# 第三步：通过表单选中读取数据
# # 1、读取内容
# #第5行，第2列
# value=sheet.cell(row=5,column=2).value
# print(value)
# # 2、写入内容
# sheet.cell(row=7,column=2,value='密码错误')
# # 写入内容一定要保存才会生效
# workbook.save('case.xlsx')
# # 3、获取最大行和最大列
# # 最大行
# max_row=sheet.max_row
# # 最大列
# max_column=sheet.max_column
# print(max_row,max_column)
# # 注意点：不要随便在表格中敲入空格
# 4、rows:按行获取所有的格子对象，每一个的格子放在一个元祖中
print(list(sheet.rows))
